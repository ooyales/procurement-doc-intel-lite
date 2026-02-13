import io
import csv
import logging
from datetime import datetime, timezone

from flask import Blueprint, request, jsonify, Response
from flask_jwt_extended import jwt_required

from app.extensions import db
from app.models.line_item import LineItem
from app.models.document import Document
from app.models.field_mapping import FieldMapping
from app.errors import BadRequestError, NotFoundError

logger = logging.getLogger(__name__)

line_items_bp = Blueprint('line_items', __name__, url_prefix='/api/line-items')


def _build_line_items_query(args):
    """Build a filtered query for line items based on request args."""
    query = LineItem.query.join(Document, LineItem.document_id == Document.id)
    query = query.filter(LineItem.session_id == '__default__')

    # Text search on product_name and part_number
    search = args.get('search', '').strip()
    if search:
        search_filter = f'%{search}%'
        query = query.filter(
            db.or_(
                LineItem.product_name.ilike(search_filter),
                LineItem.part_number.ilike(search_filter),
                LineItem.manufacturer.ilike(search_filter),
                LineItem.product_description.ilike(search_filter),
            )
        )

    # Vendor filter (join to Document)
    vendor = args.get('vendor', '').strip()
    if vendor:
        query = query.filter(Document.vendor_name.ilike(f'%{vendor}%'))

    # Category filters
    category = args.get('category', '').strip()
    if category:
        query = query.filter(LineItem.category == category)

    sub_category = args.get('sub_category', '').strip()
    if sub_category:
        query = query.filter(LineItem.sub_category == sub_category)

    # Price range
    min_price = args.get('min_price', type=float)
    if min_price is not None:
        query = query.filter(LineItem.unit_price >= min_price)

    max_price = args.get('max_price', type=float)
    if max_price is not None:
        query = query.filter(LineItem.unit_price <= max_price)

    # Date range (from Document.document_date)
    date_from = args.get('date_from', '').strip()
    if date_from:
        query = query.filter(Document.document_date >= date_from)

    date_to = args.get('date_to', '').strip()
    if date_to:
        query = query.filter(Document.document_date <= date_to)

    # Document type filter
    document_type = args.get('document_type', '').strip()
    if document_type:
        query = query.filter(Document.document_type == document_type)

    # Contract number filter
    contract_number = args.get('contract_number', '').strip()
    if contract_number:
        query = query.filter(Document.contract_number.ilike(f'%{contract_number}%'))

    # Sorting
    sort_by = args.get('sort_by', 'created_at')
    sort_order = args.get('sort_order', 'desc')

    # Map sort_by to actual column
    sort_columns = {
        'created_at': LineItem.created_at,
        'product_name': LineItem.product_name,
        'unit_price': LineItem.unit_price,
        'extended_price': LineItem.extended_price,
        'quantity': LineItem.quantity,
        'category': LineItem.category,
        'part_number': LineItem.part_number,
        'line_number': LineItem.line_number,
    }
    sort_col = sort_columns.get(sort_by, LineItem.created_at)
    if sort_order == 'asc':
        query = query.order_by(sort_col.asc())
    else:
        query = query.order_by(sort_col.desc())

    return query


@line_items_bp.route('', methods=['GET'])
@jwt_required()
def list_line_items():
    """Search and filter line items across all documents."""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 25, type=int)
    per_page = min(per_page, 100)

    query = _build_line_items_query(request.args)
    total = query.count()
    items = query.offset((page - 1) * per_page).limit(per_page).all()

    return jsonify({
        'items': [item.to_dict() for item in items],
        'total': total,
        'page': page,
        'per_page': per_page,
    })


@line_items_bp.route('/export', methods=['GET'])
@jwt_required()
def export_line_items():
    """Export filtered line items as CSV or XLSX."""
    fmt = request.args.get('format', 'csv').lower()
    if fmt not in ('csv', 'xlsx'):
        raise BadRequestError('Format must be csv or xlsx')

    query = _build_line_items_query(request.args)
    items = query.all()

    headers = [
        'Line #', 'CLIN', 'Part Number', 'Manufacturer', 'Product Name',
        'Description', 'Category', 'Sub-Category', 'Quantity', 'UOI',
        'Unit Price', 'Extended Price', 'Vendor', 'Document #',
        'Document Type', 'Document Date', 'Contract #',
    ]

    def _row(item):
        return [
            item.line_number or '',
            item.clin or '',
            item.part_number or '',
            item.manufacturer or '',
            item.product_name or '',
            item.product_description or '',
            item.category or '',
            item.sub_category or '',
            item.quantity or '',
            item.unit_of_issue or '',
            item.unit_price or '',
            item.extended_price or '',
            item.document.vendor_name if item.document else '',
            item.document.document_number if item.document else '',
            item.document.document_type if item.document else '',
            item.document.document_date if item.document else '',
            item.document.contract_number if item.document else '',
        ]

    if fmt == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for item in items:
            writer.writerow(_row(item))
        csv_data = output.getvalue()
        return Response(
            csv_data,
            mimetype='text/csv',
            headers={
                'Content-Disposition': 'attachment; filename=line_items_export.csv',
            },
        )
    else:
        # XLSX export using openpyxl
        try:
            import openpyxl
        except ImportError:
            raise BadRequestError('XLSX export requires openpyxl. Install with: pip install openpyxl')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Line Items'

        # Write headers with bold styling
        from openpyxl.styles import Font
        bold_font = Font(bold=True)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = bold_font

        # Write data rows
        for row_idx, item in enumerate(items, 2):
            for col_idx, value in enumerate(_row(item), 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit column widths (approximate)
        for col_idx, header in enumerate(headers, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(header) + 2, 12)

        xlsx_output = io.BytesIO()
        wb.save(xlsx_output)
        xlsx_output.seek(0)

        return Response(
            xlsx_output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': 'attachment; filename=line_items_export.xlsx',
            },
        )


@line_items_bp.route('/spend-analysis', methods=['GET'])
@jwt_required()
def spend_analysis():
    """Return aggregated spend analysis data."""
    base_query = LineItem.query.join(Document, LineItem.document_id == Document.id)\
        .filter(LineItem.session_id == '__default__')

    # Spend by vendor
    spend_by_vendor_rows = db.session.query(
        Document.vendor_name,
        db.func.sum(LineItem.extended_price),
    ).join(LineItem, LineItem.document_id == Document.id)\
     .filter(LineItem.session_id == '__default__')\
     .filter(Document.vendor_name.isnot(None))\
     .group_by(Document.vendor_name)\
     .order_by(db.func.sum(LineItem.extended_price).desc())\
     .all()

    spend_by_vendor = [
        {'vendor': row[0], 'total': round(row[1] or 0, 2)}
        for row in spend_by_vendor_rows
    ]

    # Spend by category
    spend_by_category_rows = db.session.query(
        LineItem.category,
        db.func.sum(LineItem.extended_price),
    ).filter(LineItem.session_id == '__default__')\
     .filter(LineItem.category.isnot(None))\
     .group_by(LineItem.category)\
     .order_by(db.func.sum(LineItem.extended_price).desc())\
     .all()

    spend_by_category = [
        {'category': row[0], 'total': round(row[1] or 0, 2)}
        for row in spend_by_category_rows
    ]

    # Spend over time (grouped by month from Document.document_date)
    # document_date is stored as string (YYYY-MM-DD or similar), so we substring for month
    spend_over_time_rows = db.session.query(
        db.func.substr(Document.document_date, 1, 7).label('month'),
        db.func.sum(LineItem.extended_price),
    ).join(LineItem, LineItem.document_id == Document.id)\
     .filter(LineItem.session_id == '__default__')\
     .filter(Document.document_date.isnot(None))\
     .filter(Document.document_date != '')\
     .group_by(db.func.substr(Document.document_date, 1, 7))\
     .order_by(db.func.substr(Document.document_date, 1, 7))\
     .all()

    spend_over_time = [
        {'month': row[0], 'total': round(row[1] or 0, 2)}
        for row in spend_over_time_rows
    ]

    return jsonify({
        'spend_by_vendor': spend_by_vendor,
        'spend_by_category': spend_by_category,
        'spend_over_time': spend_over_time,
    })


@line_items_bp.route('/<item_id>', methods=['PUT'])
@jwt_required()
def update_line_item(item_id):
    """Update an individual line item. Also updates field_mappings if a field was corrected."""
    item = LineItem.query.get(item_id)
    if not item:
        raise NotFoundError(f'Line item {item_id} not found')

    data = request.get_json()
    if not data:
        raise BadRequestError('Request body is required')

    updatable_fields = [
        'line_number', 'clin', 'slin', 'part_number', 'manufacturer',
        'manufacturer_part_number', 'product_name', 'product_description',
        'category', 'sub_category', 'quantity', 'unit_of_issue',
        'unit_price', 'extended_price', 'discount_percent', 'discount_amount',
        'labor_category', 'labor_hours', 'labor_rate',
        'period_start', 'period_end', 'human_verified',
    ]

    # Track which fields were changed for field_mapping updates
    changed_fields = []

    for field in updatable_fields:
        if field in data:
            old_value = getattr(item, field)
            new_value = data[field]
            if old_value != new_value:
                setattr(item, field, new_value)
                changed_fields.append(field)

    # If fields were corrected, update field_mappings for learning
    if changed_fields and item.document:
        vendor = item.document.vendor_name
        if vendor:
            for field in changed_fields:
                # Upsert field mapping: if a mapping exists for this vendor + target_field, update it
                existing = FieldMapping.query.filter_by(
                    vendor_name=vendor,
                    target_field=field,
                    session_id='__default__',
                ).first()

                if existing:
                    existing.times_confirmed = (existing.times_confirmed or 0) + 1
                    existing.confidence = min(1.0, (existing.confidence or 0.5) + 0.05)
                else:
                    # Create new mapping with the source column name derived from the field
                    mapping = FieldMapping(
                        vendor_name=vendor,
                        source_column_name=field,
                        target_field=field,
                        confidence=0.9,
                        times_confirmed=1,
                        session_id='__default__',
                    )
                    db.session.add(mapping)

    db.session.commit()
    return jsonify(item.to_dict())
